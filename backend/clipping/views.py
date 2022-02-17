from distutils.command import check
import os
import json
import datetime 
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from clipping.serializers import GroupKeywordSerializer, GroupScheduleSerializer, GroupSerializer, GroupUserSerializer
from utils.api import APIView, validate_serializer
from .models import Keyword, KeywordGroup, Group, GroupKeyword, GroupSchedule, GroupUser
from rest_framework.parsers import JSONParser
from openpyxl import load_workbook

def base(request):
    '''
    general page
    '''
    if request.method == 'GET':
        '''
        general page
        '''
        groups = Group.objects.all().values()
        keywords = KeywordGroup.objects.all().values()

        #========================================================#
        # Search ALL Keyword Groups for display                  #
        #========================================================#
        keyword_list = []
        for keyword in keywords:
            keyword_list.append(keyword["groupname"])

        values = {
            'groups': groups,
            'keywords': keyword_list,
            'first_depth' : 'NEWS 클리핑',
            'second_depth': 'NEWS 클리핑',
        }

        return render(request, 'clipping/clipping.html', values)
    else:
        type = request.POST['type']
        if type == 'receiver_download':
            '''
            export to excel (receiver download)
            '''
            group_id = request.POST.get('group_id', None) #그룹 id
            try:
                group = Group.objects.filter(id=group_id).first()
            except:
                return JsonResponse(data={"success":False, "data": "Clipping Group does not exist"})
            group_name = getattr(group, "name")

            wb = openpyxl.Workbook()
            sheet = wb.active

            #========================================================#
            # Search Group User list of this group                   #
            #========================================================#
            users = GroupUser.objects.filter(group_id=group_id).values()
            user_list = []
            email_list = []

            for user in users:
                user_list.append(user["name"])
                email_list.append(user["email"])

            #========================================================#
            # Make User list Excel sheet                             #
            #========================================================#
            i=1
            for ii, name in enumerate(user_list):
                sheet["A"+str(i)] = name
                sheet["B"+str(i)] = email_list[ii]
                i+=1

            filename = "%s USER LIST.xlsx" % (group_name)
            response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
        elif type == 'keyword_download':
            '''
            export to excel (keyword download)
            '''
            print("export to excel (keyword download)")
            try:
                keyword_groups = KeywordGroup.objects.all()
                keywords = Keyword.objects.all()
                if keyword_groups.exists():
                    keyword_table = {}
                    for keyword_group in keyword_groups:
                        keyword_table[keyword_group.groupname] = []
                    for keyword in keywords:
                        keyword_table[keyword.keywordgroup.groupname].append(keyword.keyword)
                    max_column = max([len(keywords) for _, keywords in keyword_table.items()])
                    print(keyword_table.items())
                    new_wb = openpyxl.Workbook()
                    new_ws = new_wb.active
                    new_ws.cell(1, 1, '키워드')
                    for col in range(2, max_column + 1):
                        new_ws.cell(1, col, f'동의어{col - 1}')
                    new_ws.cell(1, max_column + 1, '종류')
                    for idx1, keyword_group in enumerate(keyword_groups):
                        groupname = keyword_group.groupname
                        new_ws.cell(idx1 + 2, 1, groupname)
                        for idx2, synonym in enumerate(keyword_table[groupname][1:]):
                            new_ws.cell(idx1 + 2, idx2 + 2, synonym)
                        new_ws.cell(idx1 + 2, max_column + 1, keyword_group.type)
                    filename = 'Keyword.xlsx'
                    file_response = HttpResponse(
                        save_virtual_workbook(new_wb),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    file_response['Content-Disposition'] = f'attachment;filename*=UTF-8\'\'{filename}'
                    return file_response
                else:
                    return request.success()
            except Exception as e:
                return request.error("Keyword Group does not exist")


def preview(request):
    '''
    preview page
    '''
    keywords = KeywordGroup.objects.all().values()

    keyword_list = []
    for keyword in keywords:
        keyword_list.append(keyword["groupname"])
    values = {
        'keywords': keyword_list,
        'first_depth' : 'NEWS 클리핑',
        'second_depth': '미리보기',
    }
    return render(request, 'clipping/preview.html', values)

class KeywordAPI(APIView):
    def get(self, request):
        '''
        Keyword list read API
        '''
        try:
            keyword_groups = KeywordGroup.objects.all()
            keywords = Keyword.objects.all()
            if keyword_groups.exists():
                keyword_table = {}
                for keyword_group in keyword_groups:
                    keyword_table[keyword_group.groupname] = []
                for keyword in keywords:
                    keyword_table[keyword.keywordgroup.groupname].append(keyword.keyword)
                max_column = max([len(keywords) for _, keywords in keyword_table.items()])
                print(keyword_table.items())
                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.cell(1, 1, '키워드')
                for col in range(2, max_column + 1):
                    new_ws.cell(1, col, f'동의어{col - 1}')
                new_ws.cell(1, max_column + 1, '종류')
                for idx1, keyword_group in enumerate(keyword_groups):
                    groupname = keyword_group.groupname
                    new_ws.cell(idx1 + 2, 1, groupname)
                    for idx2, synonym in enumerate(keyword_table[groupname][1:]):
                        new_ws.cell(idx1 + 2, idx2 + 2, synonym)
                    new_ws.cell(idx1 + 2, max_column + 1, keyword_group.type)
                filename = 'Keyword.xlsx'
                file_response = HttpResponse(
                    save_virtual_workbook(new_wb),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                file_response['Content-Disposition'] = f'attachment;filename*=UTF-8\'\'{filename}'
                return file_response
            else:
                return self.success()
        except Exception as e:
            return self.error("Keyword Group does not exist")

    def post(self, request):
        '''
        Keyword list Excel Import
        '''
        keywordfile = request.FILES['KeywordFile']
        keywordfile = openpyxl.load_workbook(keywordfile)
        worksheet = keywordfile[keywordfile.sheetnames[0]]
        column_max = 1
        while True:
            x = worksheet.cell(row=1, column=column_max).value
            column_max += 1
            if x is None:
                break
        row = 2
        keyword_table = {}
        while True:
            keyword_groupname = worksheet.cell(row=row, column=1).value
            if keyword_groupname is None:
                break
            keyword_table[keyword_groupname] = [keyword_groupname]
            for col in range(2, column_max):
                keyword = worksheet.cell(row=row, column=col).value
                if keyword is not None:
                    keyword_table[keyword_groupname].append(keyword)
            row += 1
        Keyword.objects.all().delete()
        KeywordGroup.objects.all().delete()
        for keyword_group in keyword_table:
            group_item = KeywordGroup.objects.create(
                groupname = keyword_group,
                type = keyword_table[keyword_group][-1],
                news_platform = True,
                sns_platform = True
            )
            for keyword in keyword_table[keyword_group][:-1]:
                Keyword.objects.create(
                    keywordgroup = group_item,
                    keyword = keyword
                )
        return self.success(keyword_table)


# CLippingGroupAPI
# CRUD NEWS Clipping Groups by using 
# | get: Read Clipping Groups
# | post: Create/Update Clipping Groups
# | delete: Delete Clipping Groups
# Author: 최영우, cyw7515@naver.com
class ClippingGroupAPI(APIView):
    def get(self, request):
        '''
        Read Clipping Group API
        REQUEST FORMAT:
        {
            "group_id": group_id (num)
        }
        RESPONSE FORMAT:
        {
            "name": group_name(str),
            "collect_date": 수집기간(T:당일, F:어제) (boolean),
            "checked_keywords": 선택한 키워드 목록 (str list),
            "schedules": 수집시기 (datetime list),
        }
        '''
        group_id = request.GET.get("group_id")
        if not group_id:
            return self.error()
        
        try:
            group = Group.objects.filter(id=group_id).first()
        except:
            return self.error("Clipping Group does not exist")
        
        #========================================================#
        # Search Group Keyword List for this group               #
        #========================================================#
        check_list = []
        checked_keywords = GroupKeyword.objects.filter(group_id=group_id).values()
        for key in checked_keywords:
            check_list.append(key["keyword"])

        #========================================================#
        # Search Group schedule List for this group               #
        #========================================================#
        schedule_list = []
        schedules = GroupSchedule.objects.filter(group_id = group_id).values()
        for key in schedules:
            schedule_list.append(key["time"])

        #========================================================#
        # Make Response data                                     #
        #========================================================#
        if group:
            res_data = {}
            res_data["name"] = getattr(group, "name")
            res_data["collect_date"] = getattr(group, "collect_date")
            # data["total_keywords"] = total_keywords
            res_data["checked_keywords"] = check_list
            res_data["schedule"] = schedule_list
            return JsonResponse(data={"success":True, "data": res_data})
        else:
            return self.error("Request does not have any group name")
        
    
    def post(self, request):
        '''
        Create/Update Clipping Group API
        REQUEST FORMAT:
        {
            "user": group에 포함된 user list excel 파일 (file),
            "body":{
                "name": group name (str),
                "collect_data": 그룹 수집 기간(F: 어제, T: 당일) (boolean),
                "keywords":선택한 키워드 목록 (str list),
                "schedules": 수집 시기 (datetime list)
            } ==> string으로 전달됨, json.loads(data["body"]) 필수
        }
        RESPONSE FORMAT:
        {
            "success": True/False (boolean)
        }
        '''
        data = request.POST
        print(data)
        create_user_flag = False

        # 'not in data' means 'in FILES', so there is an attached file
        if 'users' not in data:
            file = request.FILES['users']
            create_user_flag = True
        
        data = json.loads(data['body'])
        # data = JSONParser(data)
        
        try:
            # Create Clipping Group
            group_data = {
                "name": data["name"],
                "collect_date": data["collect_date"]
            }
            exist_data = Group.objects.filter(name=data["name"]).first()
            #========================================================#
            # Create new Group if there are not same name group      #
            #========================================================#
            if exist_data is None:
                group = GroupSerializer(data=group_data)
                if group.is_valid():
                    group.save()
                else:
                    return self.error("Create clipping group data is not valid")
            #========================================================#
            # Update Group data if there is a same name group        #
            #========================================================#
            else:
                # new_data = GroupSerializer(exist_data).data
                group = GroupSerializer(exist_data, data=group_data)
                if group.is_valid():
                    group.save()
                    print(data["collect_date"])
                else:
                    return self.error("Update clipping group data is not valid")
        except:
            return self.error("cannot create Clipping Group")
        
        group_id = group.data["id"]
        
        # Create Clipping Group Users
        if create_user_flag:
            try:
                #========================================================#
                # Load Excel for external User List...                   #
                #========================================================#
                
                load_wb = load_workbook(file, data_only=True)
                load_ws = load_wb['Sheet1']
                
                user_list = []

                for row in load_ws.rows:
                    # Excel Format should be name, email, name, email...
                    user_tuple = []
                    for cell in row:
                        print(cell.value)
                        # [[name, email], [name, email], ...]
                        user_tuple.append(cell.value)
                    user_list.append(user_tuple)
                #========================================================#

                # To reflect add, update, remove user... delete all in this group
                GroupUser.objects.filter(group_id=group_id).delete()
                
                #========================================================#
                # Create Group Users in this Group                       #
                #========================================================#
                for user in user_list:
                    group_user_data = {
                        "group": group_id,
                        "name": user[0],
                        "email": user[1]
                    }
                    group_user = GroupUserSerializer(data=group_user_data)
                    if group_user.is_valid():
                        group_user.save()
                    else:
                        return self.error("Create group user data is not valid")
                #========================================================#
            except:
                return self.error("cannot create Clipping Group users")
        
        # Create Clipping Group Keyword
        try:
            # To reflect add, update, remove keyword... delete all in this group
            GroupKeyword.objects.filter(group_id=group_id).delete()

            #========================================================#
            # Create Group Keywords in this Group                    #
            #========================================================#
            for keyword in data["keywords"]:
                # print("pass key: " + keyword)
                group_keyword_data = {
                    "group": group_id,
                    "keyword": keyword
                }
                group_keyword = GroupKeywordSerializer(data=group_keyword_data)
                if group_keyword.is_valid():
                    group_keyword.save()
                else:
                    return self.error("Create group keyword data is not valid")
            #========================================================#
        except:
            return self.error("cannot create Clipping Group keywords")
        
        # Create Clipping Group Schedule
        try:
            # To reflect add, update, remove schedule... delete all in this group
            GroupSchedule.objects.filter(group_id=group_id).delete()

            #========================================================#
            # Create Group Schedules in this Group                   #
            #========================================================#
            for schedule in data["schedules"]:
                # print("pass sch: " + schedule)
                hour = int(schedule[0:2])
                min = int(schedule[3:5])
                group_schedule_data = {
                    "group": group_id,
                    "time": datetime.time(hour, min, 0)
                }
                group_schedule = GroupScheduleSerializer(data=group_schedule_data)
                if group_schedule.is_valid():
                    group_schedule.save()
                else:
                    return self.error("Create group schedule data is not valid")
            #========================================================#
        except:
            return self.error("cannot create Clipping Group schedules")

        return JsonResponse(data={"success":True})

    def delete(self, request):
        """
        Delete Clipping Group API
        REQUEST FORMAT:
        {
            "group_id": 그룹 아이디 (num)
        }
        RESPONSE FORMAT:
        {
            "success": True/False (boolean)
        }
        """
        group_id = request.GET.get("group_id")
        
        if group_id is None:
            return self.error("Group id does not exist")
        
        Group.objects.filter(id=group_id).delete()
        #Group keyword, user, schedule is deleted automatically by CASCADE

        return JsonResponse(data={"success":True})